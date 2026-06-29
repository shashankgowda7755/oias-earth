#!/usr/bin/env python3
"""
workbook-to-import — turn the canonical "COMMUNITREE Forest Trees Upload" workbook
(tabs: Sponsor Sample, Forest Sample, Tree Sample, Gifter Sample, Species Master)
into normalized JSON our APIs ingest:
  { sponsor, forest, species[], trees[] }   where trees[] is bulk_tree_gift_sheet shape.

Key transforms:
- Tree tab uses ID RANGES ("AA001-AA100") with ONE lat/lng per batch (a surveyed
  box anchor). We expand the range and spread the N trees on a ~1 ft grid around
  the anchor via WGS84 destination-point geodesy — cm-accurate, honest (positions
  inside a box are grid-modeled from the real box point, not invented GPS).
- Gifter rows merge onto trees by tree_unique_id.
- Species pass through by master_plant_species_id (Species Master = our id space).

Usage:
  python3 workbook-to-import.py <file.xlsx> [--out out.json] [--spacing-ft 1.0]
  (no --out => dry-run: prints a summary, writes nothing)
"""
import sys, re, json, math, argparse
import openpyxl

R = 6378137.0  # WGS84 mean radius (m)

def dest(lat, lng, brg, dist):
    br, la1, lo1, dr = math.radians(brg), math.radians(lat), math.radians(lng), dist / R
    la2 = math.asin(math.sin(la1) * math.cos(dr) + math.cos(la1) * math.sin(dr) * math.cos(br))
    lo2 = lo1 + math.atan2(math.sin(br) * math.sin(dr) * math.cos(la1),
                           math.cos(dr) - math.sin(la1) * math.sin(la2))
    return math.degrees(la2), math.degrees(lo2)

def spread(lat, lng, n, sp_ft):
    """N points on a near-square grid, sp_ft apart, anchored at (lat,lng) going S/E."""
    sp = sp_ft * 0.3048
    cols = max(1, math.ceil(math.sqrt(n)))
    pts = []
    for i in range(n):
        r, c = divmod(i, cols)
        plat, plng = dest(lat, lng, 180, r * sp)   # south
        plat, plng = dest(plat, plng, 90, c * sp)  # east
        pts.append((round(plat, 7), round(plng, 7)))
    return pts

def expand_range(s):
    s = str(s).strip()
    if "-" not in s:
        return [s]
    a, b = s.split("-", 1)
    ma, mb = re.match(r"^(.*?)(\d+)$", a.strip()), re.match(r"^(.*?)(\d+)$", b.strip())
    if not ma or not mb:
        return [s]
    pref = ma.group(1) or mb.group(1)
    width = len(ma.group(2))
    return [f"{pref}{str(i).zfill(width)}" for i in range(int(ma.group(2)), int(mb.group(2)) + 1)]

def slug(name):
    return re.sub(r"[^A-Z0-9]", "", str(name).upper())[:12] or "FOREST"

def rows_of(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hdr = [("" if c is None else str(c)).strip() for c in rows[0]]
    out = []
    for r in rows[1:]:
        d = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]}
        if any(v is not None and str(v).strip() for v in d.values()):
            out.append(d)
    return hdr, out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out")
    ap.add_argument("--spacing-ft", type=float, default=1.0)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    name_of = {s.lower(): s for s in wb.sheetnames}
    def tab(key):
        for k, real in name_of.items():
            if key in k:
                return wb[real]
        return None

    sp_ws, fo_ws, tr_ws, gi_ws, spm_ws = (tab("sponsor"), tab("forest"), tab("tree"), tab("gifter"), tab("species"))

    sponsor = rows_of(sp_ws)[1][0] if sp_ws and rows_of(sp_ws)[1] else None
    forest = rows_of(fo_ws)[1][0] if fo_ws and rows_of(fo_ws)[1] else None
    species = rows_of(spm_ws)[1] if spm_ws else []
    gifters = rows_of(gi_ws)[1] if gi_ws else []
    tree_rows = rows_of(tr_ws)[1] if tr_ws else []

    fuid = slug(forest.get("forest_name")) if forest else "FOREST"
    gift_by_uid = {str(g.get("tree_unique_id", "")).strip(): g for g in gifters if g.get("tree_unique_id")}

    trees = []
    skipped = []
    for row in tree_rows:
        uids = expand_range(row.get("tree_unique_id", ""))
        lat, lng = row.get("forest_tree_geo_lat"), row.get("forest_tree_geo_long")
        if lat is None or lng is None:
            skipped.append(row.get("tree_unique_id"))
            continue
        pts = spread(float(lat), float(lng), len(uids), args.spacing_ft)
        for uid, (plat, plng) in zip(uids, pts):
            g = gift_by_uid.get(uid)
            t = {
                "forest_unique_id": fuid,
                "tree_unique_id": uid,
                "species_id": int(row["master_plant_species_id"]) if row.get("master_plant_species_id") else None,
                "height": row.get("forest_tree_height"),
                "dia": row.get("forest_tree_dia"),
                "planted_on": str(row.get("planted_on"))[:10] if row.get("planted_on") else None,
                "lat": plat,
                "lng": plng,
            }
            if g:
                t["gift_recipient_name"] = g.get("name")
                t["gift_recipient_email_id"] = g.get("email_id")
            trees.append(t)

    payload = {"sponsor": sponsor, "forest": forest, "forest_unique_id": fuid,
               "species_count": len(species), "trees": trees}

    print(f"sponsor: {sponsor.get('sponsor_name') if sponsor else None}")
    print(f"forest:  {forest.get('forest_name') if forest else None}  -> unique_id {fuid}  "
          f"@ {forest.get('forest_geo_lat') if forest else '?'},{forest.get('forest_geo_long') if forest else '?'}")
    print(f"species master rows: {len(species)}")
    print(f"tree batches: {len(tree_rows)}  -> expanded trees: {len(trees)}  | gifters: {len(gifters)} matched: "
          f"{sum(1 for t in trees if t.get('gift_recipient_name'))}")
    if skipped:
        print(f"skipped batches (no coords): {skipped}")
    print("sample tree rows:")
    for t in trees[:3]:
        print("  ", json.dumps(t))

    if args.out:
        json.dump(payload, open(args.out, "w"))
        print(f"\nwrote {args.out} ({len(trees)} trees)")
    else:
        print("\n(dry-run — pass --out <file.json> to write)")

if __name__ == "__main__":
    main()
